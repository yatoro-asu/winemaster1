from collections import defaultdict
from pathlib import Path
from typing import Any, Dict, Iterable, List

from openpyxl import load_workbook

BASE_DIR = Path(__file__).resolve().parent
DATA_DIR = BASE_DIR / 'data'
DATA_DIR.mkdir(exist_ok=True)

CATEGORY_HEADERS = {'категория', 'category', 'categories'}


def normalize_value(value: Any) -> Any:
    if value is None:
        return ''
    if isinstance(value, float) and value != value:  # NaN
        return ''
    if isinstance(value, float) and value.is_integer():
        return int(value)
    return value


def read_excel_rows(path: Path) -> List[Dict[str, Any]]:
    workbook = load_workbook(filename=path, read_only=True, data_only=True)
    worksheet = workbook.active

    rows = worksheet.iter_rows(values_only=True)
    try:
        headers = [str(cell).strip() if cell is not None else '' for cell in next(rows)]
    except StopIteration:
        return []

    return [
        {
            header: normalize_value(cell)
            for header, cell in zip(headers, row)
            if header
        }
        for row in rows
        if any(cell is not None for cell in row)
    ]


def find_category_key(headers: Iterable[str]) -> str:
    lower_headers = {header.lower(): header for header in headers}
    for candidate in CATEGORY_HEADERS:
        if candidate in lower_headers:
            return lower_headers[candidate]
    raise ValueError('Category column not found in Excel headers.')


def group_by_category(records: List[Dict[str, Any]]) -> Dict[str, List[Dict[str, Any]]]:
    if not records:
        return {}

    category_key = find_category_key(records[0].keys())
    grouped: Dict[str, List[Dict[str, Any]]] = defaultdict(list)
    for record in records:
        category = normalize_value(record.get(category_key, '')) or 'Без категории'
        grouped[category].append(record)

    return grouped


def load_product_categories(path: Path) -> Dict[str, List[Dict[str, Any]]]:
    if not path.exists():
        raise FileNotFoundError(
            f'Excel file not found: {path}. '
            'Поместите файл данных в указанную папку и запустите проект снова.'
        )

    wines = read_excel_rows(path)
    return group_by_category(wines)
